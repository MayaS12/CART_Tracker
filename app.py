"""
NexCAR19 Patient Tracker — prototype
Structured replacement for the WhatsApp update workflow. Entry form is
stage-based/dynamic — which panels show depends on which stage of the CAR-T
journey the entry is for.

Run:
    pip install -r requirements.txt
    python app.py
Then open http://localhost:5000

Demo logins (change these before any real use):
    nurse1 / nurse123      (role: nurse_mo)
    spec1  / spec123       (role: specialist)
    admin  / admin123      (role: admin)
"""
import json
import io
import os
import uuid
from datetime import date, datetime

from flask import (
    Flask, render_template, redirect, url_for, request, flash, abort,
    send_file
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import RequestEntityTooLarge

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-only-change-me")
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25MB per request

# AI-drafted review suggestions — OFF unless an API key is explicitly set.
# This sends patient data (diagnosis, labs, vitals, notes) to Anthropic's
# API. Do not set this in production until Dr. Jain / Tata Memorial have
# actually signed off on that, per the DPDP Act discussion — this is not a
# decision the software should make on its own by just having a key present
# vs not, but the key gate is at least a deliberate switch, not a default.
db_url = os.environ.get("DATABASE_URL", "sqlite:///tracker.db")
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"

# ---------------- Stage / form constants ----------------

PHASES = [
    "Before Apheresis",
    "After Apheresis",
    "Conditioning",
    "Pre-Infusion & CAR T-Cell Infusion",
    "Routine Visit",
]

# Urgent-symptom keyword matching (specialist queue only) — checked against
# the free-text Symptoms box (case-insensitive substring match) and, where
# there's an objective vital to corroborate it, against the recorded vitals
# too. This is deliberately simple pattern matching, not clinical NLP — it
# will miss phrasings not on the list and can false-positive on unrelated
# text containing the same words. Treat it as a net that catches the common
# ways people phrase these, not a guarantee.
URGENT_SYMPTOM_KEYWORDS = {
    "Fever (\u2265100\u00b0F)": ["fever", "febrile", "high temp", "high fever"],
    "Difficulty breathing / shortness of breath": [
        "difficulty breathing", "trouble breathing", "shortness of breath",
        "short of breath", "breathless", "dyspnea", "dyspnoea", "hard to breathe",
        "can't breathe", "cannot breathe", "sob",
    ],
    "Low blood pressure": ["low blood pressure", "low bp", "hypotension", "hypotensive"],
    "Confusion or disorientation": ["confusion", "confused", "disoriented", "disorientation"],
    "Difficulty speaking or writing": [
        "difficulty speaking", "trouble speaking", "difficulty writing", "trouble writing",
        "slurred speech", "slurring", "aphasia", "dysarthria",
    ],
    "Vomiting or diarrhea": ["vomit", "vomiting", "diarrhea", "diarrhoea", "loose stool", "loose stools"],
    "Seizure or abnormal movements": [
        "seizure", "convulsion", "convulsing", "abnormal movement", "jerking",
        "shaking uncontrollably", "fit (medical)",
    ],
    "Severe headache": [
        "severe headache", "worst headache", "intense headache",
        "unbearable headache", "excruciating headache", "sudden headache",
    ],
    "Sudden weakness or loss of coordination": [
        "sudden weakness", "loss of coordination", "unsteady", "ataxia",
        "weakness on one side", "can't move", "cannot move", "clumsiness",
    ],
    "Bleeding": ["bleeding", "blood in stool", "blood in urine", "hemorrhage", "haemorrhage", "nosebleed", "coughing blood"],
}


def entry_urgent_flags(e):
    """Returns a list of matched urgent-symptom labels for this entry, per
    the specialist-defined checklist. Free-text keyword match against the
    Symptoms box, plus vitals-based corroboration for the three signs that
    have an objective threshold (fever, low BP, low SpO2 as a breathing-
    difficulty proxy). Returns [] if nothing matched."""
    note = (e.symptoms_note or "").lower()
    matched = []
    for label, keywords in URGENT_SYMPTOM_KEYWORDS.items():
        if any(kw in note for kw in keywords):
            matched.append(label)

    fever_label = "Fever (\u2265100\u00b0F)"
    if fever_label not in matched and e.temp_f is not None and e.temp_f >= 100.0:
        matched.append(fever_label)

    bp_label = "Low blood pressure"
    if bp_label not in matched and e.bp_sys is not None and e.bp_sys < 90:
        matched.append(bp_label)

    breathing_label = "Difficulty breathing / shortness of breath"
    if breathing_label not in matched and e.spo2 is not None and e.spo2 < 92:
        matched.append(breathing_label)

    return matched


# --- Labs: full CBC + Biochemistry panels (matching the lab's own report
# layout/order), plus the extra groups the toxicity planner needs. Field
# keys that feed computed grades (CRS/ICANS/CAR-HEMATOTOX) are preserved:
# plt, anc, hb, crp, ferritin.
LAB_GROUPS = {
    "CBC": [
        "mcv", "mch", "mchc", "rdw", "plt", "mpv", "pdw", "pct", "hb", "rbc",
        "hct", "tlc_wbc", "neutrophil_pct", "lymphocyte_pct", "monocyte_pct",
        "eosinophils_pct", "basophils_pct", "nrbcs_pct", "anc", "alc",
        "monocyte_abs", "eosinophils_abs", "basophils_abs", "ig_abs", "ig_pct",
        "w_abn_bl",
    ],
    "Biochemistry": [
        "fasting_glucose", "urea", "uric_acid", "creat", "na", "k",
        "chlorides", "bicarbonates", "serum_protein", "albumin", "globulin",
        "alp", "t_bil", "direct_bilirubin", "indirect_bilirubin", "ast", "alt",
        "ca", "po4", "mg",
    ],
    "Coagulation / inflammation": ["fibrinogen", "crp", "ldh", "ferritin", "b2_microglobulin"],
    "Immunoglobulins": ["igg", "iga", "igm"],
    "Infection workup": [
        "cmv_dna_pcr", "bdg", "galactomannan", "posa_level",
        "blood_culture", "picc_swab_culture", "sputum_culture",
    ],
    "Conditioning workup": ["triglycerides", "beta_hcg"],
}
LAB_FIELDS = [f for group in LAB_GROUPS.values() for f in group]

LAB_LABELS = {
    "mcv": "MCV", "mch": "MCH", "mchc": "MCHC", "rdw": "RDW", "plt": "Platelets",
    "mpv": "MPV", "pdw": "PDW", "pct": "PCT", "hb": "HB", "rbc": "RBC", "hct": "HCT",
    "tlc_wbc": "TLC", "neutrophil_pct": "Neutrophil %", "lymphocyte_pct": "Lymphocyte %",
    "monocyte_pct": "Monocyte %", "eosinophils_pct": "Eosinophils %",
    "basophils_pct": "Basophils %", "nrbcs_pct": "NRBCS %",
    "anc": "Neutrophils abs (ANC)", "alc": "Lymphocyte abs (ALC)",
    "monocyte_abs": "Monocyte abs", "eosinophils_abs": "Eosinophils abs",
    "basophils_abs": "Basophils abs", "ig_abs": "IG ABS", "ig_pct": "IG %",
    "w_abn_bl": "W_ABN_BL (Blasts/Abn Lympho)",
    "fasting_glucose": "Fasting Plasma Glucose", "urea": "Serum Urea",
    "uric_acid": "Serum Uric Acid", "creat": "Serum Creatinine",
    "na": "Serum Sodium", "k": "Serum Potassium", "chlorides": "Serum Chlorides",
    "bicarbonates": "Serum Bicarbonates", "serum_protein": "Serum Protein",
    "albumin": "Serum Albumin", "globulin": "Serum Globulin",
    "alp": "Serum Alkaline Phosphatase", "t_bil": "Serum Total Bilirubin",
    "direct_bilirubin": "Serum Direct Bilirubin", "indirect_bilirubin": "Serum Indirect Bilirubin",
    "ast": "Serum AST", "alt": "Serum ALT", "ca": "Serum Calcium",
    "po4": "Serum Phosphorus", "mg": "Serum Magnesium",
}

# Reference range shown next to each field (no computed High/Low flag — just
# the range itself, per the lab's own report). Ranges for the CBC and
# Biochemistry panels are taken directly from the reports sent; the rest are
# unit-only since no specific range was supplied for them — treat those as
# illustrative, not lab-validated, and confirm with Dr. Jain before relying
# on them clinically.
LAB_REFERENCE = {
    "mcv": "83-101 fL", "mch": "27-32 pg", "mchc": "31.5-34.5 g/dL", "rdw": "11.6-14.0 %CV",
    "plt": "150-400 x10\u2079/L", "mpv": "7.5-10.5 fL", "pdw": "25-65 %", "pct": "0.12-0.36 %",
    "hb": "13.0-17.0 g/dL", "rbc": "4.5-5.5 x10\u00b9\u00b2/L", "hct": "40-50 %",
    "tlc_wbc": "4.0-10.0 x10\u2079/L", "neutrophil_pct": "40.0-80.0 %", "lymphocyte_pct": "20.0-40.0 %",
    "monocyte_pct": "2.0-10.0 %", "eosinophils_pct": "1.0-6.0 %", "basophils_pct": "1.0-2.0 %",
    "nrbcs_pct": "0.0-2.0 NRBCS/100", "anc": "2.0-7.0 x10\u2079/L", "alc": "1.0-3.0 x10\u2079/L",
    "monocyte_abs": "0.2-1.0 x10\u2079/L", "eosinophils_abs": "0.02-0.5 x10\u2079/L",
    "basophils_abs": "0.02-0.1 x10\u2079/L", "ig_abs": "0.00-5.00 x10\u2079/L", "ig_pct": "0.00-5.00 %",
    "w_abn_bl": "qualitative \u2014 flag if present",
    "fasting_glucose": "70-99 mg/dL", "urea": "12.84-42.8 mg/dL", "uric_acid": "3.5-7.2 mg/dL",
    "creat": "0.7-1.3 mg/dL", "na": "136-145 mmol/L", "k": "3.5-5.1 mmol/L",
    "chlorides": "98-107 mmol/L", "bicarbonates": "22-29 mmol/L", "serum_protein": "6.6-8.3 g/dL",
    "albumin": "3.5-5.2 g/dL", "globulin": "1.7-3.5 g/dL", "alp": "30-120 U/L",
    "t_bil": "0.3-1.2 mg/dL", "direct_bilirubin": "0.0-0.2 mg/dL", "indirect_bilirubin": "0.3-1.0 mg/dL",
    "ast": "< 50 U/L", "alt": "< 50 U/L", "ca": "8.6-10 mg/dL", "po4": "2.7-4.5 mg/dL",
    "mg": "1.8-2.6 mg/dL",
    "fibrinogen": "mg/dL", "crp": "mg/dL", "ldh": "U/L", "ferritin": "ng/mL", "b2_microglobulin": "mg/L",
    "igg": "mg/dL", "iga": "mg/dL", "igm": "mg/dL",
    "triglycerides": "mg/dL", "beta_hcg": "mIU/mL",
}

# Panels/lab-groups that only apply to specific stages — everything else
# (Vitals, Symptoms, the core lab groups, meds, notes) is shown on every stage.
PHASE_ONLY_LAB_GROUPS = {"Conditioning workup": "Conditioning"}

ICAHT_GRADES = ["none", "Grade I", "Grade II", "Grade III", "Grade IV"]
BM_MRD_OPTIONS = ["not_done", "negative", "positive"]
CSF_STATUS_OPTIONS = ["not_done", "negative", "flow_positive", "morpho_positive", "both_positive"]

DRUGS_STOPPED_OPTIONS = [
    "steroids", "immunosuppressants", "antiproliferative_therapies",
    "tyrosine_kinase_inhibitors", "hydroxyurea", "cytotoxic_drugs",
    "weekly_maintenance_therapies", "anti_cd20_antibodies",
    "cns_disease_prophylaxis", "radiation_therapy",
]
CONSUMABLES_OPTIONS = [
    "apheresis_kit", "acd", "saline", "heparin",
    "machine_functioning_confirmed", "investigations_noted",
    "cd3_count_informed_to_manufacturing",
]
CONDITIONING_INVESTIGATIONS_OPTIONS = [
    "cbc", "biochemistry", "serum_electrolytes", "s_crp", "s_ferritin",
    "s_beta2_microglobulin", "ldh", "s_triglycerides", "fibrinogen",
    "beta_hcg", "pet_ct", "bm_biopsy",
]


# ---------------- Models ----------------

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(128))
    role = db.Column(db.String(32), nullable=False)
    active = db.Column(db.Boolean, default=True, nullable=False)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)


class Patient(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128))
    code = db.Column(db.String(32), unique=True, nullable=False)
    age = db.Column(db.Integer)
    sex = db.Column(db.String(16))

    treating_center = db.Column(db.String(128))
    referring_center = db.Column(db.String(128))

    diagnosis = db.Column(db.String(256))            # "Baseline diagnosis"
    diagnosis_subtype = db.Column(db.String(128))
    cytogenetics_baseline = db.Column(db.String(256))
    molecular_report = db.Column(db.Text)
    indication_for_cart = db.Column(db.Text)
    relapse_date = db.Column(db.Date, nullable=True)
    disease_history = db.Column(db.Text)

    disease_status = db.Column(db.String(32), nullable=True)  # "B-cell ALL" / "B-cell lymphomas"

    all_blast_pct = db.Column(db.Float, nullable=True)
    all_mrd_assessment = db.Column(db.String(128), nullable=True)
    all_csf_assessment_done = db.Column(db.String(8), nullable=True)
    all_csf_involved = db.Column(db.String(8), nullable=True)

    lymphoma_disease_stage = db.Column(db.String(64), nullable=True)
    lymphoma_supradiaphragmatic_nodes = db.Column(db.Text, nullable=True)
    lymphoma_infradiaphragmatic_nodes = db.Column(db.Text, nullable=True)
    lymphoma_extranodal_disease = db.Column(db.String(8), nullable=True)
    lymphoma_bulky_disease = db.Column(db.String(8), nullable=True)
    lymphoma_bulky_size = db.Column(db.String(64), nullable=True)
    lymphoma_bm_involvement = db.Column(db.String(8), nullable=True)
    lymphoma_ldh = db.Column(db.Float, nullable=True)

    prior_therapy_lines = db.Column(db.Text)  # JSON list of {start, stop, drugs_dose, response}

    car_t_product = db.Column(db.String(128))
    planned_dose = db.Column(db.String(64))
    central_line_type = db.Column(db.String(64))
    central_line_date = db.Column(db.Date, nullable=True)

    infusion_date = db.Column(db.Date, nullable=True)

    entries = db.relationship("Entry", backref="patient", lazy=True,
                               order_by="Entry.entry_date")

    def prior_therapy_list(self):
        return json.loads(self.prior_therapy_lines or "[]")

    def current_status(self):
        """The operational 'at a glance' picture for the top of the patient
        page — where things stand right now, useful to a nurse or specialist
        alike the moment they open the chart. Uses the single most recent
        entry regardless of review status (current vitals/meds shouldn't
        wait on approval to be visible), plus counts of this patient's own
        entries sitting in the review queue or needing revision."""
        if not self.entries:
            return None
        latest = max(self.entries, key=lambda e: e.entry_date)
        active_meds = [m for m in latest.med_rows()
                       if not m.get("stop_date") or m["stop_date"] >= date.today().isoformat()]
        pending_count = sum(1 for e in self.entries if e.status == "submitted")
        revision_count = sum(1 for e in self.entries if e.status == "changes_requested")
        return {
            "latest_entry": latest, "active_meds": active_meds,
            "pending_count": pending_count, "revision_count": revision_count,
        }

    def latest_toxicities(self):
        approved = [e for e in sorted(self.entries, key=lambda e: e.entry_date, reverse=True)
                    if e.status == "approved"]
        result = {"crs_grade": None, "icans_grade": None, "hlh_suspected": None,
                  "igg": None, "infection_findings": None, "fever_present": False, "as_of": None}
        for e in approved:
            if result["crs_grade"] is None and e.crs_grade:
                result["crs_grade"] = e.crs_grade
                result["as_of"] = result["as_of"] or e.entry_date
            if result["icans_grade"] is None and e.icans_grade:
                result["icans_grade"] = e.icans_grade
                result["as_of"] = result["as_of"] or e.entry_date
            if result["hlh_suspected"] is None and e.hlh_suspected:
                result["hlh_suspected"] = e.hlh_suspected
            if not result["fever_present"] and e.fever_present:
                result["fever_present"] = True
            if result["igg"] is None:
                igg = e.lab_dict().get("igg")
                if igg:
                    result["igg"] = igg
            if result["infection_findings"] is None:
                if e.infection_note:
                    result["infection_findings"] = e.infection_note
                    result["as_of"] = result["as_of"] or e.entry_date
                else:
                    ld = e.lab_dict()
                    bits = [f"{k}: {ld[k]}" for k in
                            ("blood_culture", "picc_swab_culture", "sputum_culture") if ld.get(k)]
                    if bits:
                        result["infection_findings"] = "; ".join(bits)
        return result


class Entry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey("patient.id"), nullable=False)
    submission_token = db.Column(db.String(64), nullable=True, unique=True)

    phase = db.Column(db.String(64))
    entry_date = db.Column(db.Date, default=date.today)
    reason_for_admission = db.Column(db.String(256))
    performance_status = db.Column(db.Integer)

    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    status = db.Column(db.String(32), default="submitted")
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    reviewed_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    reviewed_at = db.Column(db.DateTime, nullable=True)
    review_comment = db.Column(db.Text, nullable=True)

    temp_f = db.Column(db.Float)
    pulse = db.Column(db.Integer)
    bp_sys = db.Column(db.Integer)
    bp_dia = db.Column(db.Integer)
    spo2 = db.Column(db.Integer)
    rr = db.Column(db.Integer)

    symptom_flags = db.Column(db.Text)
    symptoms_note = db.Column(db.Text)

    labs = db.Column(db.Text)

    assessment_date = db.Column(db.Date, nullable=True)
    bm_blast_pct = db.Column(db.Float, nullable=True)
    bm_mrd_status = db.Column(db.String(16), nullable=True)
    csf_status = db.Column(db.String(16), nullable=True)
    ctg_findings = db.Column(db.String(256), nullable=True)
    rt_qpcr = db.Column(db.String(256), nullable=True)
    molecular_report = db.Column(db.String(256), nullable=True)

    apheresis_fit = db.Column(db.String(8), nullable=True)
    apheresis_labs_wnl = db.Column(db.String(8), nullable=True)
    meal_before_apheresis = db.Column(db.String(8), nullable=True)
    heparin_issued = db.Column(db.String(8), nullable=True)
    apheresis_machine_ok = db.Column(db.String(8), nullable=True)
    apheresis_concerns = db.Column(db.Text, nullable=True)
    apheresis_final_fit = db.Column(db.String(8), nullable=True)
    non_refundable_blocked = db.Column(db.String(8), nullable=True)
    drugs_stopped = db.Column(db.Text, nullable=True)
    consumables_ready = db.Column(db.Text, nullable=True)
    pending_reports = db.Column(db.Text, nullable=True)
    apheresis_time = db.Column(db.String(64), nullable=True)
    apheresis_outcome = db.Column(db.String(16), nullable=True)

    bridging_therapy_planned = db.Column(db.String(8), nullable=True)
    it_mtx_given = db.Column(db.String(8), nullable=True)
    it_mtx_dose = db.Column(db.String(64), nullable=True)
    response_assessment_done = db.Column(db.String(8), nullable=True)
    response_assessment_type = db.Column(db.String(64), nullable=True)
    response_assessment_notes = db.Column(db.Text, nullable=True)
    response_assessment_date = db.Column(db.Date, nullable=True)
    central_line_removal_precautions = db.Column(db.String(8), nullable=True)

    conditioning_investigations_done = db.Column(db.Text, nullable=True)
    car_hematotox_score = db.Column(db.Integer, nullable=True)
    car_hematotox_risk = db.Column(db.String(32), nullable=True)
    prn_orders_confirmed = db.Column(db.String(8), nullable=True)
    bsa = db.Column(db.Float, nullable=True)
    fludarabine_given = db.Column(db.String(8), nullable=True)
    fludarabine_dose = db.Column(db.String(64), nullable=True)
    cyclophosphamide_given = db.Column(db.String(8), nullable=True)
    cyclophosphamide_dose = db.Column(db.String(64), nullable=True)
    dose_modification = db.Column(db.String(8), nullable=True)
    dose_modification_reason = db.Column(db.Text, nullable=True)

    ps_unchanged = db.Column(db.String(8), nullable=True)
    chemo_toxicities_resolved = db.Column(db.String(8), nullable=True)
    postponement_pulmonary = db.Column(db.String(8), nullable=True)
    postponement_cardiac = db.Column(db.String(8), nullable=True)
    postponement_hypotension = db.Column(db.String(8), nullable=True)
    postponement_infection = db.Column(db.String(8), nullable=True)
    fit_for_infusion = db.Column(db.String(8), nullable=True)
    delay_reassessment_date = db.Column(db.Date, nullable=True)
    delay_duration = db.Column(db.String(64), nullable=True)
    delay_reason = db.Column(db.String(256), nullable=True)

    car_t_dose_given = db.Column(db.String(64), nullable=True)
    premedication_start = db.Column(db.String(32), nullable=True)
    premedication_stop = db.Column(db.String(32), nullable=True)
    thaw_start = db.Column(db.String(32), nullable=True)
    thaw_stop = db.Column(db.String(32), nullable=True)
    infusion_start = db.Column(db.String(32), nullable=True)
    infusion_stop = db.Column(db.String(32), nullable=True)
    infusion_outcome = db.Column(db.String(16), nullable=True)

    fever_present = db.Column(db.Boolean, default=False)
    hypotension_level = db.Column(db.String(32), default="none")
    hypoxia_level = db.Column(db.String(32), default="none")
    crs_grade = db.Column(db.Integer, default=0)

    ice_orientation = db.Column(db.Integer, default=4)
    ice_naming = db.Column(db.Integer, default=3)
    ice_command = db.Column(db.Integer, default=1)
    ice_writing = db.Column(db.Integer, default=1)
    ice_attention = db.Column(db.Integer, default=1)
    consciousness_level = db.Column(db.String(32), default="spontaneous")
    seizure_status = db.Column(db.String(32), default="none")
    motor_findings = db.Column(db.String(32), default="none")
    cerebral_edema = db.Column(db.String(32), default="none")
    ice_score = db.Column(db.Integer)
    icans_grade = db.Column(db.Integer, default=0)

    icaht_grade = db.Column(db.String(16), nullable=True)
    hlh_suspected = db.Column(db.String(8), nullable=True)
    hlh_note = db.Column(db.Text, nullable=True)
    infection_note = db.Column(db.Text, nullable=True)
    other_toxicity_note = db.Column(db.Text, nullable=True)

    medication_rows = db.Column(db.Text)
    medications_plan = db.Column(db.Text)
    free_notes = db.Column(db.Text)

    creator = db.relationship("User", foreign_keys=[created_by_id])
    reviewer = db.relationship("User", foreign_keys=[reviewed_by_id])

    def symptom_list(self):
        return json.loads(self.symptom_flags or "[]")

    def lab_dict(self):
        return json.loads(self.labs or "{}")

    def med_rows(self):
        return json.loads(self.medication_rows or "[]")

    def drugs_stopped_list(self):
        return json.loads(self.drugs_stopped or "[]")

    def consumables_list(self):
        return json.loads(self.consumables_ready or "[]")

    def conditioning_investigations_list(self):
        return json.loads(self.conditioning_investigations_done or "[]")

    def day_relative(self):
        if self.patient.infusion_date and self.entry_date:
            delta = (self.entry_date - self.patient.infusion_date).days
            return f"D{'+' if delta >= 0 else ''}{delta}"
        return "—"

    def summary_text(self):
        """A short deterministic narrative of this visit — not AI-generated,
        just the entry's own fields assembled into readable sentences, so
        it's always available with no API dependency and nothing invented."""
        parts = []

        header = f"{self.phase} on {self.entry_date}"
        if self.performance_status is not None:
            header += f", PS {self.performance_status}"
        parts.append(header + ".")

        if any(v is not None for v in (self.temp_f, self.pulse, self.bp_sys, self.spo2)):
            parts.append(
                f"Vitals: T {self.temp_f if self.temp_f is not None else '—'}\u00b0F, "
                f"P {self.pulse if self.pulse is not None else '—'}, "
                f"BP {self.bp_sys if self.bp_sys is not None else '—'}/"
                f"{self.bp_dia if self.bp_dia is not None else '—'}, "
                f"SpO2 {self.spo2 if self.spo2 is not None else '—'}%."
            )
        if self.symptoms_note:
            parts.append(f"Symptoms: {self.symptoms_note}.")

        if self.crs_grade or self.icans_grade:
            parts.append(f"CRS G{self.crs_grade or 0}, ICANS G{self.icans_grade or 0} "
                          f"(ICE {self.ice_score if self.ice_score is not None else '—'}/10).")
        if self.icaht_grade and self.icaht_grade != "none":
            parts.append(f"ICAHT {self.icaht_grade}.")
        if self.hlh_suspected == "yes":
            parts.append("HLH-like syndrome suspected.")
        if self.infection_note:
            parts.append(f"Infections: {self.infection_note}.")
        if self.other_toxicity_note:
            parts.append(f"Other: {self.other_toxicity_note}.")
        if self.car_hematotox_score is not None and self.phase == "Conditioning":
            parts.append(f"CAR-HEMATOTOX {self.car_hematotox_score} points ({self.car_hematotox_risk}).")

        if self.phase == "Before Apheresis" and self.apheresis_outcome:
            parts.append(f"Apheresis: {self.apheresis_outcome}.")
        if self.phase == "After Apheresis" and self.response_assessment_done == "yes":
            parts.append(f"Response assessment: {self.response_assessment_type or 'done'}"
                          + (f" — {self.response_assessment_notes}" if self.response_assessment_notes else "") + ".")
        if self.phase == "Pre-Infusion & CAR T-Cell Infusion":
            if self.fit_for_infusion:
                parts.append(f"Fit for infusion: {self.fit_for_infusion}.")
            if self.infusion_outcome:
                parts.append(f"Infusion: {self.infusion_outcome}.")
        if self.phase == "Routine Visit" and self.assessment_date:
            bits = []
            if self.bm_blast_pct is not None:
                bits.append(f"BM blasts {self.bm_blast_pct}%")
            if self.bm_mrd_status:
                bits.append(f"MRD {self.bm_mrd_status}")
            if self.csf_status:
                bits.append(f"CSF {self.csf_status}")
            if bits:
                parts.append("Disease reassessment: " + ", ".join(bits) + ".")

        meds = self.med_rows()
        if meds:
            parts.append("Meds: " + ", ".join(m["name"] for m in meds) + ".")
        if self.medications_plan:
            parts.append(f"Plan: {self.medications_plan}.")
        elif self.free_notes:
            parts.append(f"Notes: {self.free_notes}.")

        return " ".join(parts)


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    entry_id = db.Column(db.Integer, db.ForeignKey("entry.id"))
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    action = db.Column(db.String(64))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    detail = db.Column(db.Text)

    entry = db.relationship("Entry")
    user = db.relationship("User")


class Attachment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    entry_id = db.Column(db.Integer, db.ForeignKey("entry.id"), nullable=True)
    patient_id = db.Column(db.Integer, db.ForeignKey("patient.id"), nullable=True)
    category = db.Column(db.String(64), nullable=True)
    filename = db.Column(db.String(256))
    content_type = db.Column(db.String(128))
    data = db.Column(db.LargeBinary)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

    entry = db.relationship("Entry", backref=db.backref("attachments", lazy=True))
    patient = db.relationship("Patient", backref=db.backref("attachments", lazy=True))
    uploader = db.relationship("User")

    def is_image(self):
        return (self.content_type or "").startswith("image/")

    def is_video(self):
        return (self.content_type or "").startswith("video/")


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ---------------- Computed grading ----------------

def compute_crs_grade(fever, hypotension_level, hypoxia_level):
    if not fever:
        return 0
    grade = 1
    if hypotension_level == "multi_vasopressor" or hypoxia_level == "positive_pressure":
        grade = 4
    elif hypotension_level == "one_vasopressor" or hypoxia_level == "high_flow_or_mask":
        grade = 3
    elif hypotension_level == "fluids" or hypoxia_level == "low_flow":
        grade = 2
    return grade


def compute_icans(ice_score, consciousness, seizure, motor, edema):
    if consciousness == "unarousable":
        grade = 4
    else:
        if ice_score >= 7:
            grade = 1
        elif ice_score >= 3:
            grade = 2
        else:
            grade = 3

    if consciousness == "voice":
        grade = max(grade, 2)
    elif consciousness == "tactile":
        grade = max(grade, 3)

    if seizure == "resolves_lt5min":
        grade = max(grade, 3)
    elif seizure == "prolonged_or_recurrent":
        grade = max(grade, 4)

    if motor == "focal_weakness":
        grade = max(grade, 4)

    if edema == "focal":
        grade = max(grade, 3)
    elif edema == "diffuse":
        grade = max(grade, 4)

    return grade


def compute_hematotox(labs):
    def num(key):
        v = labs.get(key)
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    score = 0
    plt, anc, hb, crp, ferritin = num("plt"), num("anc"), num("hb"), num("crp"), num("ferritin")

    if plt is not None:
        score += 2 if plt <= 75 else (1 if plt <= 175 else 0)
    if anc is not None:
        score += 1 if anc < 1.2 else 0
    if hb is not None:
        score += 1 if hb <= 9.0 else 0
    if crp is not None:
        score += 1 if crp >= 3.0 else 0
    if ferritin is not None:
        score += 2 if ferritin > 2000 else (1 if ferritin >= 650 else 0)

    risk = "High risk (HThigh)" if score >= 2 else "Low risk (HTlow)"
    return score, risk




# ---------------- Auth / dashboard routes ----------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"]).first()
        if user and not user.active:
            flash("This account has been deactivated")
        elif user and user.check_password(request.form["password"]):
            login_user(user)
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    q = request.args.get("q", "").strip()
    patients_query = Patient.query
    if q:
        like = f"%{q}%"
        if current_user.role == "nurse_mo":
            patients_query = patients_query.filter(
                db.or_(Patient.name.ilike(like), Patient.code.ilike(like))
            )
        else:
            patients_query = patients_query.filter(
                db.or_(
                    Patient.name.ilike(like), Patient.code.ilike(like),
                    Patient.treating_center.ilike(like), Patient.referring_center.ilike(like),
                )
            )
    patients = patients_query.order_by(Patient.code).all()
    pending_count = Entry.query.filter_by(status="submitted").count()
    revision_count = Entry.query.filter_by(status="changes_requested").count()
    return render_template("dashboard.html", patients=patients,
                            pending_count=pending_count, revision_count=revision_count, q=q)


def _parse_date(val):
    return datetime.strptime(val, "%Y-%m-%d").date() if val else None


ALLOWED_ATTACHMENT_TYPES = ("image/", "video/")
ALLOWED_REPORT_TYPES = ("image/", "video/", "application/pdf")


def save_attachments(files, entry=None, patient=None, category=None, allowed_types=ALLOWED_ATTACHMENT_TYPES):
    saved, skipped = 0, []
    for f in files:
        if not f or not f.filename:
            continue
        content_type = f.content_type or ""
        if not content_type.startswith(allowed_types):
            skipped.append(f.filename)
            continue
        db.session.add(Attachment(
            entry_id=entry.id if entry else None,
            patient_id=patient.id if patient else None,
            category=category, filename=f.filename, content_type=content_type,
            data=f.read(), uploaded_by_id=current_user.id,
        ))
        saved += 1
    return saved, skipped


@app.route("/patients/new", methods=["GET", "POST"])
@login_required
def new_patient():
    if request.method == "POST":
        f = request.form

        def opt_float(key):
            v = f.get(key)
            return float(v) if v else None

        prior_lines = []
        starts = f.getlist("line_start")
        stops = f.getlist("line_stop")
        drugs = f.getlist("line_drugs_dose")
        responses = f.getlist("line_response")
        for i in range(len(starts)):
            if starts[i] or drugs[i] or responses[i]:
                prior_lines.append({
                    "start": starts[i], "stop": stops[i] if i < len(stops) else "",
                    "drugs_dose": drugs[i] if i < len(drugs) else "",
                    "response": responses[i] if i < len(responses) else "",
                })

        p = Patient(
            name=f.get("name"),
            code=f["code"],
            age=f.get("age") or None,
            sex=f.get("sex"),
            treating_center=f.get("treating_center"),
            referring_center=f.get("referring_center"),
            diagnosis=f.get("diagnosis"),
            diagnosis_subtype=f.get("diagnosis_subtype"),
            cytogenetics_baseline=f.get("cytogenetics_baseline"),
            molecular_report=f.get("molecular_report"),
            indication_for_cart=f.get("indication_for_cart"),
            relapse_date=_parse_date(f.get("relapse_date")),
            disease_history=f.get("disease_history"),
            disease_status=f.get("disease_status") or None,
            all_blast_pct=opt_float("all_blast_pct"),
            all_mrd_assessment=f.get("all_mrd_assessment"),
            all_csf_assessment_done=f.get("all_csf_assessment_done") or None,
            all_csf_involved=f.get("all_csf_involved") or None,
            lymphoma_disease_stage=f.get("lymphoma_disease_stage"),
            lymphoma_supradiaphragmatic_nodes=f.get("lymphoma_supradiaphragmatic_nodes"),
            lymphoma_infradiaphragmatic_nodes=f.get("lymphoma_infradiaphragmatic_nodes"),
            lymphoma_extranodal_disease=f.get("lymphoma_extranodal_disease") or None,
            lymphoma_bulky_disease=f.get("lymphoma_bulky_disease") or None,
            lymphoma_bulky_size=f.get("lymphoma_bulky_size"),
            lymphoma_bm_involvement=f.get("lymphoma_bm_involvement") or None,
            lymphoma_ldh=opt_float("lymphoma_ldh"),
            prior_therapy_lines=json.dumps(prior_lines),
            car_t_product=f.get("car_t_product"),
            planned_dose=f.get("planned_dose"),
            central_line_type=f.get("central_line_type"),
            central_line_date=_parse_date(f.get("central_line_date")),
            infusion_date=_parse_date(f.get("infusion_date")),
        )
        db.session.add(p)
        db.session.commit()

        saved, skipped = save_attachments(
            request.files.getlist("disease_reports"), patient=p,
            category="diagnosis_report", allowed_types=ALLOWED_REPORT_TYPES,
        )
        db.session.commit()
        if skipped:
            flash(f"Skipped {len(skipped)} file(s) that weren't images/videos/PDFs: {', '.join(skipped)}")
        flash(f"Patient {p.code} created")
        return redirect(url_for("dashboard"))
    return render_template("new_patient.html")


@app.route("/patients/<int:patient_id>")
@login_required
def patient_timeline(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    entries = Entry.query.filter_by(patient_id=patient_id).order_by(Entry.entry_date).all()
    return render_template("timeline.html", patient=patient, entries=entries,
                            toxicities=patient.latest_toxicities(),
                            status=patient.current_status())


# ---------------- Entry create / edit (shared logic) ----------------

def parse_med_rows(f):
    names = f.getlist("med_name")
    freqs = f.getlist("med_freq")
    doses = f.getlist("med_dose")
    starts = f.getlist("med_start")
    stops = f.getlist("med_stop")
    rows = []
    for i, name in enumerate(names):
        if name.strip():
            rows.append({
                "name": name.strip(),
                "frequency": freqs[i].strip() if i < len(freqs) else "",
                "dose": doses[i].strip() if i < len(doses) else "",
                "start_date": starts[i].strip() if i < len(starts) else "",
                "stop_date": stops[i].strip() if i < len(stops) else "",
            })
    return rows


def apply_entry_fields(e, f):
    labs = {k: f.get(f"lab_{k}") for k in LAB_FIELDS if f.get(f"lab_{k}")}

    fever = f.get("fever_present") == "yes"
    crs_grade = compute_crs_grade(fever, f.get("hypotension_level", "none"),
                                   f.get("hypoxia_level", "none"))
    ice_parts = [int(f.get(k, 0) or 0) for k in
                 ["ice_orientation", "ice_naming", "ice_command",
                  "ice_writing", "ice_attention"]]
    ice_score = sum(ice_parts)
    icans_grade = compute_icans(ice_score, f.get("consciousness_level", "spontaneous"),
                                 f.get("seizure_status", "none"),
                                 f.get("motor_findings", "none"),
                                 f.get("cerebral_edema", "none"))
    hematotox_score, hematotox_risk = compute_hematotox(labs)

    def opt_int(key):
        v = f.get(key)
        return int(v) if v else None

    def opt_float(key):
        v = f.get(key)
        return float(v) if v else None

    e.phase = f["phase"]
    e.entry_date = _parse_date(f["entry_date"])
    e.reason_for_admission = f.get("reason_for_admission")
    e.performance_status = opt_int("performance_status")

    e.temp_f, e.pulse = opt_float("temp_f"), opt_int("pulse")
    e.bp_sys, e.bp_dia = opt_int("bp_sys"), opt_int("bp_dia")
    e.spo2, e.rr = opt_int("spo2"), opt_int("rr")

    e.symptom_flags = json.dumps([])  # checklist retired — symptoms are free text now
    e.symptoms_note = f.get("symptoms_note")

    e.labs = json.dumps(labs)

    e.assessment_date = _parse_date(f.get("assessment_date"))
    e.bm_blast_pct = opt_float("bm_blast_pct")
    e.bm_mrd_status = f.get("bm_mrd_status") or None
    e.csf_status = f.get("csf_status") or None
    e.ctg_findings = f.get("ctg_findings")
    e.rt_qpcr = f.get("rt_qpcr")
    e.molecular_report = f.get("molecular_report")

    e.apheresis_fit = f.get("apheresis_fit") or None
    e.apheresis_labs_wnl = f.get("apheresis_labs_wnl") or None
    e.meal_before_apheresis = f.get("meal_before_apheresis") or None
    e.heparin_issued = f.get("heparin_issued") or None
    e.apheresis_machine_ok = f.get("apheresis_machine_ok") or None
    e.apheresis_concerns = f.get("apheresis_concerns")
    e.apheresis_final_fit = f.get("apheresis_final_fit") or None
    e.non_refundable_blocked = f.get("non_refundable_blocked") or None
    e.drugs_stopped = json.dumps(f.getlist("drugs_stopped"))
    e.consumables_ready = json.dumps(f.getlist("consumables_ready"))
    e.pending_reports = f.get("pending_reports")
    e.apheresis_time = f.get("apheresis_time")
    e.apheresis_outcome = f.get("apheresis_outcome") or None

    e.bridging_therapy_planned = f.get("bridging_therapy_planned") or None
    e.it_mtx_given = f.get("it_mtx_given") or None
    e.it_mtx_dose = f.get("it_mtx_dose")
    e.response_assessment_done = f.get("response_assessment_done") or None
    e.response_assessment_type = f.get("response_assessment_type")
    e.response_assessment_notes = f.get("response_assessment_notes")
    e.response_assessment_date = _parse_date(f.get("response_assessment_date"))
    e.central_line_removal_precautions = f.get("central_line_removal_precautions") or None

    e.conditioning_investigations_done = json.dumps(f.getlist("conditioning_investigations_done"))
    e.car_hematotox_score = hematotox_score
    e.car_hematotox_risk = hematotox_risk
    e.prn_orders_confirmed = f.get("prn_orders_confirmed") or None
    e.bsa = opt_float("bsa")
    e.fludarabine_given = f.get("fludarabine_given") or None
    e.fludarabine_dose = f.get("fludarabine_dose")
    e.cyclophosphamide_given = f.get("cyclophosphamide_given") or None
    e.cyclophosphamide_dose = f.get("cyclophosphamide_dose")
    e.dose_modification = f.get("dose_modification") or None
    e.dose_modification_reason = f.get("dose_modification_reason")

    e.ps_unchanged = f.get("ps_unchanged") or None
    e.chemo_toxicities_resolved = f.get("chemo_toxicities_resolved") or None
    e.postponement_pulmonary = f.get("postponement_pulmonary") or None
    e.postponement_cardiac = f.get("postponement_cardiac") or None
    e.postponement_hypotension = f.get("postponement_hypotension") or None
    e.postponement_infection = f.get("postponement_infection") or None
    e.fit_for_infusion = f.get("fit_for_infusion") or None
    e.delay_reassessment_date = _parse_date(f.get("delay_reassessment_date"))
    e.delay_duration = f.get("delay_duration")
    e.delay_reason = f.get("delay_reason")

    e.car_t_dose_given = f.get("car_t_dose_given")
    e.premedication_start = f.get("premedication_start")
    e.premedication_stop = f.get("premedication_stop")
    e.thaw_start, e.thaw_stop = f.get("thaw_start"), f.get("thaw_stop")
    e.infusion_start, e.infusion_stop = f.get("infusion_start"), f.get("infusion_stop")
    e.infusion_outcome = f.get("infusion_outcome") or None

    e.fever_present = fever
    e.hypotension_level = f.get("hypotension_level", "none")
    e.hypoxia_level = f.get("hypoxia_level", "none")
    e.crs_grade = crs_grade

    e.ice_orientation, e.ice_naming = ice_parts[0], ice_parts[1]
    e.ice_command, e.ice_writing, e.ice_attention = ice_parts[2], ice_parts[3], ice_parts[4]
    e.consciousness_level = f.get("consciousness_level", "spontaneous")
    e.seizure_status = f.get("seizure_status", "none")
    e.motor_findings = f.get("motor_findings", "none")
    e.cerebral_edema = f.get("cerebral_edema", "none")
    e.ice_score, e.icans_grade = ice_score, icans_grade

    e.icaht_grade = f.get("icaht_grade") or None
    e.hlh_suspected = f.get("hlh_suspected") or None
    e.hlh_note = f.get("hlh_note")
    e.infection_note = f.get("infection_note")
    e.other_toxicity_note = f.get("other_toxicity_note")

    e.medication_rows = json.dumps(parse_med_rows(f))
    e.medications_plan = f.get("medications_plan")
    e.free_notes = f.get("free_notes")


@app.route("/patients/<int:patient_id>/entries/new", methods=["GET", "POST"])
@login_required
def new_entry(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    if request.method == "POST":
        token = request.form.get("submission_token")
        if token:
            existing = Entry.query.filter_by(submission_token=token).first()
            if existing:
                flash("That entry was already submitted")
                return redirect(url_for("patient_timeline", patient_id=patient.id))

        e = Entry(patient_id=patient.id, created_by_id=current_user.id,
                  status="submitted", submission_token=token)
        apply_entry_fields(e, request.form)
        db.session.add(e)
        db.session.commit()

        saved, skipped = save_attachments(request.files.getlist("attachments"), entry=e)
        db.session.commit()

        db.session.add(AuditLog(entry_id=e.id, user_id=current_user.id,
                                 action="submitted", detail=f"Phase: {e.phase}"))
        db.session.commit()
        if skipped:
            flash(f"Skipped {len(skipped)} file(s) that weren't images/videos: {', '.join(skipped)}")
        flash("Entry submitted for specialist review")
        return redirect(url_for("patient_timeline", patient_id=patient.id))

    return render_template(
        "entry_form.html", patient=patient, entry=None, phases=PHASES,
        lab_groups=LAB_GROUPS,
        lab_labels=LAB_LABELS, lab_reference=LAB_REFERENCE,
        phase_only_lab_groups=PHASE_ONLY_LAB_GROUPS,
        icaht_grades=ICAHT_GRADES, bm_mrd_options=BM_MRD_OPTIONS,
        csf_status_options=CSF_STATUS_OPTIONS,
        drugs_stopped_options=DRUGS_STOPPED_OPTIONS,
        consumables_options=CONSUMABLES_OPTIONS,
        conditioning_investigations_options=CONDITIONING_INVESTIGATIONS_OPTIONS,
        today=date.today().isoformat(),
        submission_token=uuid.uuid4().hex,
    )


@app.route("/entries/<int:entry_id>/edit", methods=["GET", "POST"])
@login_required
def edit_entry(entry_id):
    e = Entry.query.get_or_404(entry_id)
    if e.status != "changes_requested":
        flash("Only entries marked 'changes requested' can be edited")
        return redirect(url_for("patient_timeline", patient_id=e.patient_id))

    if request.method == "POST":
        apply_entry_fields(e, request.form)
        e.status = "submitted"
        e.submitted_at = datetime.utcnow()
        db.session.commit()

        saved, skipped = save_attachments(request.files.getlist("attachments"), entry=e)
        db.session.commit()

        db.session.add(AuditLog(entry_id=e.id, user_id=current_user.id,
                                 action="resubmitted",
                                 detail=f"Revised after review comment: {e.review_comment or ''}"))
        db.session.commit()
        if skipped:
            flash(f"Skipped {len(skipped)} file(s) that weren't images/videos: {', '.join(skipped)}")
        flash("Entry resubmitted for specialist review")
        return redirect(url_for("patient_timeline", patient_id=e.patient_id))

    return render_template(
        "entry_form.html", patient=e.patient, entry=e, phases=PHASES,
        lab_groups=LAB_GROUPS,
        lab_labels=LAB_LABELS, lab_reference=LAB_REFERENCE,
        phase_only_lab_groups=PHASE_ONLY_LAB_GROUPS,
        icaht_grades=ICAHT_GRADES, bm_mrd_options=BM_MRD_OPTIONS,
        csf_status_options=CSF_STATUS_OPTIONS,
        drugs_stopped_options=DRUGS_STOPPED_OPTIONS,
        consumables_options=CONSUMABLES_OPTIONS,
        conditioning_investigations_options=CONDITIONING_INVESTIGATIONS_OPTIONS,
        today=date.today().isoformat(),
    )


@app.route("/attachments/<int:attachment_id>")
@login_required
def view_attachment(attachment_id):
    a = Attachment.query.get_or_404(attachment_id)
    return send_file(io.BytesIO(a.data), mimetype=a.content_type,
                      download_name=a.filename, as_attachment=False)


def build_patient_summary_pdf(patient):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleX", parent=styles["Title"], fontSize=17)
    meta_style = ParagraphStyle("MetaX", parent=styles["Normal"], fontSize=10,
                                 textColor=colors.HexColor("#556070"), spaceAfter=6)
    visit_header_style = ParagraphStyle("VisitHeaderX", parent=styles["Heading3"],
                                         fontSize=11.5, textColor=colors.HexColor("#14324d"),
                                         spaceBefore=12, spaceAfter=3)
    status_style = ParagraphStyle("StatusX", parent=styles["Normal"], fontSize=8.5,
                                   textColor=colors.HexColor("#72808d"), spaceAfter=4)
    body_style = ParagraphStyle("BodyX", parent=styles["Normal"], fontSize=10, leading=14)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.7 * inch, rightMargin=0.7 * inch,
                             topMargin=0.7 * inch, bottomMargin=0.7 * inch)
    story = [
        Paragraph(patient.name or patient.code, title_style),
        Paragraph(
            f"{patient.code} &middot; {patient.diagnosis or 'diagnosis not on file'}"
            f"{f' ({patient.diagnosis_subtype})' if patient.diagnosis_subtype else ''} &middot; "
            f"Infusion date: {patient.infusion_date or 'not yet'}",
            meta_style,
        ),
        HRFlowable(width="100%", color=colors.HexColor("#dde3ea")),
    ]

    entries = sorted(patient.entries, key=lambda e: e.entry_date)
    if not entries:
        story.append(Spacer(1, 10))
        story.append(Paragraph("No visits recorded yet.", body_style))
    for e in entries:
        story.append(Paragraph(
            f"{e.entry_date} \u2014 {e.phase} ({e.day_relative()})", visit_header_style))
        story.append(Paragraph(
            f"Status: {e.status.replace('_', ' ')} &middot; Entered by {e.creator.full_name if e.creator else '—'}"
            + (f" &middot; Reviewed by {e.reviewer.full_name}" if e.reviewer else ""),
            status_style,
        ))
        story.append(Paragraph(e.summary_text(), body_style))

    doc.build(story)
    buf.seek(0)
    return buf


@app.route("/patients/<int:patient_id>/summary_pdf")
@login_required
def patient_summary_pdf(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    buf = build_patient_summary_pdf(patient)
    filename = f"{patient.code}_visit_summaries.pdf"
    return send_file(buf, mimetype="application/pdf", download_name=filename, as_attachment=True)


PATIENT_XLSX_COLUMNS = [
    "Code", "Name", "Age", "Sex", "Treating center", "Diagnosis", "Subtype",
    "Cytogenetics", "Disease status", "Relapse date", "Indication for CAR-T",
    "CAR-T product", "Planned dose", "Infusion date", "Entries",
    "Latest CRS", "Latest ICANS", "Latest HLH suspected", "Latest IgG", "Latest infections",
]
VISIT_XLSX_COLUMNS = [
    "Date", "Day rel.", "Phase", "Status", "PS", "Temp F", "Pulse", "BP Sys", "BP Dia", "SpO2",
    "Symptoms", "CRS grade", "ICANS grade", "ICE score", "ICAHT grade", "HLH suspected",
    "Infections", "Other", "Labs", "Medications", "Plan notes", "Free notes",
    "Entered by", "Reviewed by",
]


def _patient_xlsx_row(p):
    tox = p.latest_toxicities()
    return [
        p.code, p.name, p.age, p.sex, p.treating_center, p.diagnosis, p.diagnosis_subtype,
        p.cytogenetics_baseline, p.disease_status, p.relapse_date, p.indication_for_cart,
        p.car_t_product, p.planned_dose, p.infusion_date, len(p.entries),
        tox["crs_grade"], tox["icans_grade"], tox["hlh_suspected"], tox["igg"], tox["infection_findings"],
    ]


def _visit_xlsx_row(e):
    ld = e.lab_dict()
    labs_str = ", ".join(f"{k}={v}" for k, v in ld.items())
    meds_str = "; ".join(f"{m['name']} ({m['frequency']}, {m['dose']})" for m in e.med_rows())
    return [
        e.entry_date, e.day_relative(), e.phase, e.status, e.performance_status,
        e.temp_f, e.pulse, e.bp_sys, e.bp_dia, e.spo2,
        e.symptoms_note, e.crs_grade, e.icans_grade, e.ice_score, e.icaht_grade,
        e.hlh_suspected, e.infection_note, e.other_toxicity_note,
        labs_str, meds_str, e.medications_plan, e.free_notes,
        e.creator.full_name if e.creator else None,
        e.reviewer.full_name if e.reviewer else None,
    ]


def _autosize_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 40)


def build_patient_workbook(patient):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Patient"
    ws1.append(PATIENT_XLSX_COLUMNS)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    ws1.append(_patient_xlsx_row(patient))
    _autosize_columns(ws1)

    ws2 = wb.create_sheet("Visits")
    ws2.append(VISIT_XLSX_COLUMNS)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for e in sorted(patient.entries, key=lambda e: e.entry_date):
        ws2.append(_visit_xlsx_row(e))
    _autosize_columns(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_all_patients_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Patients"
    ws1.append(PATIENT_XLSX_COLUMNS)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    patients = Patient.query.order_by(Patient.code).all()
    for p in patients:
        ws1.append(_patient_xlsx_row(p))
    _autosize_columns(ws1)

    ws2 = wb.create_sheet("All Visits")
    ws2.append(["Patient code", "Patient name"] + VISIT_XLSX_COLUMNS)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for p in patients:
        for e in sorted(p.entries, key=lambda e: e.entry_date):
            ws2.append([p.code, p.name] + _visit_xlsx_row(e))
    _autosize_columns(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@app.route("/patients/<int:patient_id>/export.xlsx")
@login_required
def patient_export_excel(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    buf = build_patient_workbook(patient)
    return send_file(buf, mimetype=XLSX_MIMETYPE,
                      download_name=f"{patient.code}_export.xlsx", as_attachment=True)


@app.route("/patients/export_all.xlsx")
@login_required
def all_patients_export_excel():
    if current_user.role not in ("specialist", "admin"):
        abort(403)
    buf = build_all_patients_workbook()
    return send_file(buf, mimetype=XLSX_MIMETYPE,
                      download_name="all_patients_export.xlsx", as_attachment=True)


@app.route("/attachments/<int:attachment_id>/delete", methods=["POST"])
@login_required
def delete_attachment(attachment_id):
    a = Attachment.query.get_or_404(attachment_id)
    if a.entry and a.entry.status == "approved":
        flash("Can't remove attachments from an approved entry")
        return redirect(url_for("patient_timeline", patient_id=a.entry.patient_id))
    patient_id = a.entry.patient_id if a.entry else a.patient_id
    db.session.delete(a)
    db.session.commit()
    flash("Attachment removed")
    return redirect(request.referrer or url_for("patient_timeline", patient_id=patient_id))


@app.errorhandler(RequestEntityTooLarge)
def handle_large_upload(e):
    flash("That upload was too large — photos/videos are capped at 25MB total per submission")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/revisions")
@login_required
def revisions():
    needing_changes = Entry.query.filter_by(status="changes_requested") \
        .order_by(Entry.reviewed_at.desc()).all()
    return render_template("revisions.html", entries=needing_changes)


@app.route("/queue")
@login_required
def queue():
    if current_user.role not in ("specialist", "admin"):
        abort(403)
    pending = Entry.query.filter_by(status="submitted").order_by(Entry.submitted_at).all()
    # Flagged entries (matched one or more of the specialist-defined urgent
    # symptoms) surface first; everything else keeps the normal
    # oldest-submitted-first order.
    rows = [(e, entry_urgent_flags(e)) for e in pending]
    rows.sort(key=lambda row: (not row[1], row[0].submitted_at))
    return render_template("queue.html", entries=rows)


@app.route("/entries/<int:entry_id>/review", methods=["POST"])
@login_required
def review_entry(entry_id):
    if current_user.role not in ("specialist", "admin"):
        abort(403)
    e = Entry.query.get_or_404(entry_id)
    decision = request.form["decision"]
    e.status = "approved" if decision == "approve" else "changes_requested"
    e.reviewed_by_id = current_user.id
    e.reviewed_at = datetime.utcnow()
    e.review_comment = request.form.get("comment")
    db.session.add(AuditLog(entry_id=e.id, user_id=current_user.id,
                             action=e.status, detail=e.review_comment or ""))
    db.session.commit()
    flash(f"Entry {'approved' if decision == 'approve' else 'sent back for changes'}")
    return redirect(url_for("queue"))


@app.route("/audit")
@login_required
def audit_log():
    if current_user.role not in ("specialist", "admin"):
        abort(403)
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(300).all()
    return render_template("audit.html", logs=logs)


@app.route("/users")
@login_required
def users():
    if current_user.role != "admin":
        abort(403)
    all_users = User.query.order_by(User.username).all()
    return render_template("users.html", users=all_users)


@app.route("/users/new", methods=["GET", "POST"])
@login_required
def new_user():
    if current_user.role != "admin":
        abort(403)
    if request.method == "POST":
        f = request.form
        if User.query.filter_by(username=f["username"]).first():
            flash("That username is already taken")
            return redirect(url_for("new_user"))
        u = User(
            username=f["username"],
            full_name=f.get("full_name"),
            role=f["role"],
            password_hash=generate_password_hash(f["password"]),
        )
        db.session.add(u)
        db.session.commit()
        flash(f"Account created for {u.full_name} ({u.username})")
        return redirect(url_for("users"))
    return render_template("new_user.html")


@app.route("/users/<int:user_id>/toggle", methods=["POST"])
@login_required
def toggle_user(user_id):
    if current_user.role != "admin":
        abort(403)
    u = User.query.get_or_404(user_id)
    if u.id == current_user.id:
        flash("You can't deactivate your own account while logged in as it")
        return redirect(url_for("users"))
    u.active = not u.active
    db.session.commit()
    flash(f"{u.full_name} ({u.username}) {'reactivated' if u.active else 'deactivated'}")
    return redirect(url_for("users"))


# ---------------- bootstrap / self-healing schema ----------------

def ensure_schema():
    inspector = inspect(db.engine)
    is_sqlite = db.engine.url.get_backend_name() == "sqlite"
    type_map = {
        db.Integer: "INTEGER", db.Float: "FLOAT", db.Boolean: "BOOLEAN",
        db.Date: "DATE", db.DateTime: "TIMESTAMP", db.Text: "TEXT",
    }

    for table_name, model in (("user", User), ("patient", Patient), ("entry", Entry),
                               ("attachment", Attachment)):
        if table_name not in inspector.get_table_names():
            continue
        existing_cols = {c["name"] for c in inspector.get_columns(table_name)}
        for col in model.__table__.columns:
            if col.name in existing_cols:
                continue
            col_type = "TEXT"
            for py_type, sql_type in type_map.items():
                if isinstance(col.type, py_type):
                    col_type = sql_type
                    break
            if isinstance(col.type, db.String):
                col_type = "TEXT" if not is_sqlite else "VARCHAR"
            ident = f'"{table_name}"' if table_name == "user" else table_name
            try:
                with db.engine.connect() as conn:
                    conn.execute(text(f'ALTER TABLE {ident} ADD COLUMN {col.name} {col_type}'))
                    conn.commit()
            except Exception:
                pass


def seed():
    db.create_all()
    ensure_schema()
    if not User.query.filter_by(username="nurse1").first():
        db.session.add(User(username="nurse1", full_name="Nurse Demo",
                             role="nurse_mo",
                             password_hash=generate_password_hash("nurse123")))
    if not User.query.filter_by(username="spec1").first():
        db.session.add(User(username="spec1", full_name="Specialist Demo",
                             role="specialist",
                             password_hash=generate_password_hash("spec123")))
    if not User.query.filter_by(username="admin").first():
        db.session.add(User(username="admin", full_name="Admin Demo",
                             role="admin",
                             password_hash=generate_password_hash("admin123")))
    db.session.commit()


with app.app_context():
    seed()

if __name__ == "__main__":
    app.run(debug=True)
